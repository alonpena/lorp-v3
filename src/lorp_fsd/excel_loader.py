"""Load Excel ``LoRP-FSD`` rows into :class:`ExperimentConfig` records.

Only the ``LoRP-FSD`` sheet is read. Real depot IDs are preserved from the
``Depot1..4`` labels (``'d5' -> 5``), along with the size/capacity/demand/usage/
vehicle slots. Rows are never silently dropped for filename mismatch — instance
resolution is a separate concern (:func:`lorp_fsd.dat_parser.resolve_dat_path`).

``problemID`` is not an Excel column; the benchmark rows are ``problemID=0``
(Arslan), confirmed by the C congruency test, so it is injected as 0.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Union

import openpyxl

from .experiment_config import ExperimentConfig, SelectedDepot

PathLike = Union[str, Path]

SHEET_NAME = "LoRP-FSD"
MAX_DEPOT_SLOTS = 4


def _to_float(v) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    return float(v)


def _to_int(v) -> Optional[int]:
    f = _to_float(v)
    return None if f is None else int(round(f))


def _parse_depot_label(label) -> Optional[int]:
    """'d5' -> 5, 5 -> 5, '' / None -> None."""
    if label is None:
        return None
    if isinstance(label, (int, float)):
        return int(label)
    s = str(label).strip().lower()
    if not s:
        return None
    if s.startswith("d"):
        s = s[1:]
    if not s:
        return None
    return int(float(s))


def load_lorp_fsd_rows(xlsx_path: PathLike) -> List[ExperimentConfig]:
    """Read every data row of the ``LoRP-FSD`` sheet into an ExperimentConfig."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"workbook has no sheet {SHEET_NAME!r}; sheets={wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    configs: List[ExperimentConfig] = []
    for r_index, row in enumerate(rows):
        name = get(row, "name")
        if name is None or (isinstance(name, str) and not name.strip()):
            continue  # blank trailing row, not a filename mismatch

        selected: Dict[int, SelectedDepot] = {}
        for k in range(1, MAX_DEPOT_SLOTS + 1):
            depot_id = _parse_depot_label(get(row, f"Depot{k}"))
            if depot_id is None:
                continue
            selected[depot_id] = SelectedDepot(
                depot_id=depot_id,
                size=_to_int(get(row, f"sizeD{k}")) or 0,
                capacity=_to_float(get(row, f"CapD{k}")) or 0.0,
                demand=_to_float(get(row, f"DemandD{k}")),
                usage=_to_float(get(row, f"%UsageD{k}")),
                vehicles=_to_int(get(row, f"VehiclesD{k}")),
            )

        configs.append(
            ExperimentConfig(
                name=str(name).strip(),
                F_R=_to_float(get(row, "F_R")),
                F_A=_to_float(get(row, "F_A")),
                R=_to_float(get(row, "R")),
                Length=_to_float(get(row, "Length")),
                problem_id=0,
                of=str(get(row, "of") or "cost"),
                UB=_to_float(get(row, "UB")),
                LB=_to_float(get(row, "LB")),
                status=(str(get(row, "Status name")).strip() if get(row, "Status name") is not None else None),
                gap=_to_float(get(row, "gap")),
                cost_routing=_to_float(get(row, "Cost Routing")),
                cost_vehicles=_to_float(get(row, "Cost (Vehicles)")),
                cost_depots=_to_float(get(row, "Cost (Depots)")),
                cost_direct_all=_to_float(get(row, "Cost Direct All")),
                selected_depots=selected,
                total_depots=_to_int(get(row, "TotalDepots")),
                total_vehicles=_to_int(get(row, "TotalVehicles")),
                row_index=r_index,
            )
        )

    wb.close()
    return configs


def load_row(xlsx_path: PathLike, index: int) -> ExperimentConfig:
    """Load a single ``LoRP-FSD`` data row by 0-based index."""
    rows = load_lorp_fsd_rows(xlsx_path)
    return rows[index]


__all__ = ["load_lorp_fsd_rows", "load_row", "SHEET_NAME"]
